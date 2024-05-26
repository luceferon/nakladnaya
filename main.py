import os
import shutil
from datetime import datetime
from tkinter import messagebox
from PyQt5 import QtGui
from PyQt5.QtCore import QRegExp
from PyQt5.QtWidgets import (QApplication, QComboBox, QMessageBox)
from PyQt5.uic import loadUiType
import openpyxl
import sys
import win32api

Ui_MainWindow, QMainWindow = loadUiType("MainWindowUI.ui")


class FormFiller:
    def __init__(self):
        self.save_folder = "Save"
        self.print_folder = "Print"

    def check_and_fill_form_from_excel(self):
        files_in_save = [f for f in os.listdir(self.save_folder) if f.endswith('.xlsx')]

        if len(files_in_save) == 1:
            file_path = os.path.join(self.save_folder, files_in_save[0])
            wb = openpyxl.load_workbook(file_path)
            sheet = wb.active

            # заполняем форму из файла
            recipient = sheet["C5"].value
            names = [sheet[f"B{9 + i}"].value for i in range(1, 21)]
            quantities = [sheet[f"I{9 + i}"].value for i in range(1, 21)]

            wb.close()

            return recipient, names, quantities

        # если файла нет копируем последний распечатанный
        elif len(files_in_save) == 0:

            files_in_print = [f for f in os.listdir(self.print_folder) if f.endswith('.xlsx')]

            if len(files_in_print) > 0:
                latest_file = max(files_in_print)

                source_file_path = os.path.join(self.print_folder, latest_file)

                destination_file_path = os.path.join(self.save_folder, latest_file)

                shutil.copyfile(source_file_path, destination_file_path)

            return None, None, None


class MainWindow(QMainWindow, Ui_MainWindow):
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.PBClose.clicked.connect(self.close)
        # Очистка содержимого формы
        self.PBClear.clicked.connect(self.clearLineEditContents)
        # Сохраняем накладную
        self.PBSave.clicked.connect(self.save_data_to_excel)
        # Сохраняем накладную и печатаем
        self.PBPrint.clicked.connect(self.save_data_and_print_to_excel)
        #Заполняем дату
        self.DEDate.setDate(datetime.now().date())

        #Заполняем ответственых
        self.CBAgreed.addItems(["Начальник  участка- Крицкий С.Г.", "ИО Начальника  участка- Чащин Ю.М.",
                                "ИО Начальника  участка- Скурихин А. А."])

        #Заполняем направление куда ТМЦ
        self.CBRecipient.addItems(["Уч. Кувай", "г. Красноярск, ул. 26 Бакинских Комиссаров - Склад Б/У",
                                   "г. Красноярск, ул. 26 Бакинских Комиссаров", "Уч. Караган"])

        #Заполняем транспорт
        self.CBTransport.addItems(["MAN Г/н С615ЕР124",
                                   "Mitsubishi L200 Г/н С673ОВ124",
                                   "FUSO Г/н К700ХВ124",
                                   "MAN Г/н Х253ВО124 Прицеп Г/н МВ5912 24",
                                   "ISUZU RUSTRAC Г/н Т213СК124",
                                   "MAN Г/н С469ЕР124",
                                   "Камаз 65117-62  Г/н В775НВ124",
                                   "УАЗ Профи Г/н Н868РУ124",
                                   "НИВА Г/н Е776ТЕ124",
                                   "SCANIA Г/н С774СН 124 Прицеп Г/н МС9034 24",
                                   "MAN  Г/н М231ОЕ124",
                                   "Mitsubishi L200  Г/н Х674ОА124",
                                   "НИВА Chevrolet  Г/н  Е594ЕТ19",
                                   "Fiat Fullback Г/н У746ТЕ124",
                                   "JAC Т6  Г/н  В808ТМ124",
                                   "HINO Ranger Г/н В195КУ19",
                                   "Mitsubishi L200 Г/н А642РУ124",
                                   "ГАЗ Г/н У708УР124",
                                   "GREAT WALL POER Г/н Т891РК124",
                                   "JAC Т6  Г/н  К515ТМ124",
                                   "НИВА Г/н К709РН1244",
                                   "КАМАЗ Г/н О528ХА124",
                                   "Mitsubishi L200 Г/н К961МТ124",
                                   "МАЗ Г/н А233РЕ124",
                                   "HINO Ranger Г/н Х074КМ19",
                                   "Mitsubishi L200 Г/н К863ТВ124",
                                   "НИВА Г/н Е502ОР124",
                                   "КАМАЗ Г/н М473АК124"])

        #Заполняем водителей
        self.CBDriver.addItems(["Каменев И. В.",
                                "Рыжков С. А.",
                                "Метелкин Д. С.",
                                "Жигунов В. А.",
                                "Агапов В. А.",
                                "Арбузин И. В.",
                                "Газизов И. О.",
                                "Коробейников А. В.",
                                "Краско А. В.",
                                "Марцыновский Р. В.",
                                "Серединский М. И.",
                                "Хуторской Д. Н.",
                                "Газизов А. В.",
                                "Большаков В. С.",
                                "Лебеденко Ю. В.",
                                "Петров И. С.",
                                "Голынский В. Н.",
                                "Евдокимов В. В.",
                                "Лященко С. Н.",
                                "Татти Л. И.",
                                "Степаненко А. В.",
                                "Калашников С. С.",
                                "Даровских В. С.",
                                "Красильников А. Ю.",
                                "Степень В.В.",
                                "Щенников В. А. ",
                                "Елисейкин А. Н.",
                                "Болотов Д.Н.",
                                "Чебодаев В. А."])

        #Сортируем транспорт и водителей
        self.CBTransport.model().sort(0)
        self.CBDriver.model().sort(0)

        #Заполняем еденицу измерения ТМЦ
        for cb_piece in self.findChildren(QComboBox, QRegExp("CBPiece\\d+")):
            cb_piece.addItems([" ", "Шт.", "л.", "м2", "т.", "кг."])

        #Заполняем номер накладной исходя из того есть ли сохраненная накладная
        # либо если ее нет смотрим последнию напечатаную и прибавляем 1
        save_folder = "./Save"
        print_folder = "./Print"
        max_number = 0
        if os.path.exists(save_folder) and os.path.isdir(save_folder):
            for file in os.listdir(save_folder):
                if file.endswith(".xlsx"):
                    file_number = int(''.join(filter(str.isdigit, file)))
                    max_number = max(max_number, file_number)
        if max_number == 0:
            if os.path.exists(print_folder) and os.path.isdir(print_folder):
                for file in os.listdir(print_folder):
                    if file.endswith(".xlsx"):
                        file_number = int(''.join(filter(str.isdigit, file)))
                        max_number = max(max_number, file_number)
            max_number += 1
        self.SBNumber.setValue(max_number)

        #Проверяем есть ли сохраненная накладная и если да то загружаем ее, если нет то копируем последнию накладную
        form_filler = FormFiller()
        recipient, names, quantities = form_filler.check_and_fill_form_from_excel()

        if recipient is not None:
            self.CBRecipient.setCurrentText(recipient)
            for idx, name in enumerate(names):
                getattr(self, f"LEName{idx + 1}").setText(name)
            for idx, quantity in enumerate(quantities):
                if quantity is not None:
                    getattr(self, f"LEQuantity{idx + 1}").setText(str(quantity))

        self.printed = False

    def clearLineEditContents(self):
        for i in range(1, 21):
            le_name = getattr(self, f"LEName{i}", None)
            le_quantity = getattr(self, f"LEQuantity{i}", None)
            if le_name:
                le_name.clear()
            if le_quantity:
                le_quantity.clear()

    def save_data_to_excel(self):
        save_dir = "Save"
        if not os.path.exists(save_dir):
            os.makedirs(save_dir)

        files = [f for f in os.listdir(save_dir) if f.endswith('.xlsx')]

        if files:
            file_path = os.path.join(save_dir, files[0])
        else:
            print_dir = "Print"
            print_files = [f for f in os.listdir(print_dir) if f.endswith('.xlsx')]
            if print_files:
                latest_file = max(print_files, key=os.path.getctime)
                shutil.copy(os.path.join(print_dir, latest_file), os.path.join(save_dir, latest_file))
                file_path = os.path.join(save_dir, latest_file)
            else:
                print("No *.xlsx file found in the 'Print' directory")
                return

        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        # Записываем данные из компонентов в ячейки
        sheet['E1'] = self.SBNumber.value()

        row = 10
        for i in range(1, 21):
            name = getattr(self, f'LEName{i}').text()
            piece = getattr(self, f'CBPiece{i}').currentText()
            quantity = getattr(self, f'LEQuantity{i}').text()

            sheet[f'B{row}'] = name
            sheet[f'H{row}'] = piece
            sheet[f'I{row}'] = quantity

            row += 1
        file_name = f"{self.SBNumber.value()}.xlsx"
        file_path = os.path.join(save_dir, file_name)
        workbook.save(file_path)
        print("Data saved to:", file_path)
        for file in os.listdir(save_dir):
            if os.path.isfile(os.path.join(save_dir, file)) and file != file_name:
                os.remove(os.path.join(save_dir, file))

    def save_data_and_print_to_excel(self):
        save_dir = "Save"
        if not os.path.exists(save_dir):
            os.makedirs(save_dir)

        files = [f for f in os.listdir(save_dir) if f.endswith('.xlsx')]

        if files:
            file_path = os.path.join(save_dir, files[0])
        else:
            print_dir = "Print"
            print_files = [f for f in os.listdir(print_dir) if f.endswith('.xlsx')]
            if print_files:
                latest_file = max(print_files, key=os.path.getctime)
                shutil.copy(os.path.join(print_dir, latest_file), os.path.join(save_dir, latest_file))
                file_path = os.path.join(save_dir, latest_file)
            else:
                print("No *.xlsx file found in the 'Print' directory")
                return

        confirmation = messagebox.askyesno("Подтверждение", "Действительно распечатать накладную? "
                                                            "Данные о перевозчике верны?")

        if not confirmation:
            return

        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        # Записываем данные из компонентов в ячейки
        sheet['E1'] = "№" + str(self.SBNumber.value())
        sheet['G1'] = self.DEDate.date().toString("dd.MM.yyyy")
        sheet['G3'] = self.CBAgreed.currentText()
        sheet['C5'] = self.CBRecipient.currentText()
        sheet['D7'] = self.CBTransport.currentText()
        sheet['D35'] = self.CBDriver.currentText()

        row = 10
        for i in range(1, 21):
            name = getattr(self, f'LEName{i}').text()
            piece = getattr(self, f'CBPiece{i}').currentText()
            quantity = getattr(self, f'LEQuantity{i}').text()

            sheet[f'B{row}'] = name
            sheet[f'H{row}'] = piece
            sheet[f'I{row}'] = quantity

            row += 1
        file_name = f"{self.SBNumber.value()}.xlsx"
        print_dir = "Print"
        file_path = os.path.join(print_dir, file_name)
        workbook.save(file_path)
        print("Data saved to:", file_path)

        win32api.ShellExecute(0, "print", file_path, None, ".", 0)

        for file in os.listdir(save_dir):
            file_path = os.path.join(save_dir, file)
            if os.path.isfile(file_path):
                os.remove(file_path)
        self.printed = True
        self.close()

    def closeEvent(self, event):
        if isinstance(event, QtGui.QCloseEvent):
            if not self.printed:  # Проверяем флаг печати
                reply = QMessageBox.question(self, 'Подтверждение', 'Вы уверены, что хотите закрыть окно? Не сохраненные данные будут утеряны',
                                             QMessageBox.Yes | QMessageBox.No, QMessageBox.No)

                if reply == QMessageBox.Yes:
                    event.accept()
                else:
                    event.ignore()
            else:
                event.accept()  # Закрываем окно без запроса, если печать была завершена
        else:
            event.accept()


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())
